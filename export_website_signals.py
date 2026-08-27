"""Export the Zee Business website signal archive for a day to Excel."""
import datetime as dt
import json
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill

DAY = "2026-08-27"
API_URL = "https://www.zeebusinesslive.com/api/trades?day=0"
OUTPUT = Path("zeebiz_website_signals_2026-08-27.xlsx")

payload = requests.get(API_URL, timeout=30).json()
if payload.get("date") != DAY:
    raise RuntimeError(f"Website returned {payload.get('date')}, expected {DAY}")

def ist_time(value):
    if not value:
        return ""
    parsed = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    return parsed.astimezone(dt.timezone(dt.timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d %H:%M:%S")

def analyst(item):
    if item.get("sebiAdvisorName"):
        return item["sebiAdvisorName"]
    message = item.get("message") or ""
    known = {
        "सोमिल मेहता": "Somil Mehta",
        "सहज अग्रवाल": "Sahaj Aggarwal",
        "विकास सेठी": "Vikas Sethi",
        "विकास सालुंखे": "Vikas Salunkhe",
        "सुदीप शाह": "Sudeep Shah",
        "सिद्धार्थ सेडानी": "Siddharth Sedani",
        "Somil Mehta": "Somil Mehta",
        "Vikas Sethi": "Vikas Sethi",
    }
    for name, normalized in known.items():
        if name.lower() in message.lower():
            return normalized
    # Preserve the original message when the API could not normalize a name.
    return "(see message)" if message else ""

records = []
for item in payload.get("data", []):
    entry_time = ist_time(item.get("entryOn") or item.get("createdOn"))
    if not entry_time.startswith(DAY):
        continue
    hour, minute = map(int, entry_time[11:16].split(":"))
    if not (dt.time(9, 15) <= dt.time(hour, minute) <= dt.time(15, 30)):
        continue
    records.append({
        "Date/Time (IST)": entry_time,
        "Ticker": item.get("ticker", ""),
        "Position": item.get("position", ""),
        "Entry Price": item.get("entryPrice", ""),
        "Target": item.get("target", ""),
        "Stop Loss": item.get("stoploss", ""),
        "Analyst": analyst(item),
        "Message / Analyst Text": item.get("message", ""),
        "Category": item.get("category", ""),
        "Source": "Zee Business website API",
    })

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Website Signals"
headers = list(records[0]) if records else ["Date/Time (IST)", "Ticker"]
sheet.append(headers)
for cell in sheet[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
for row in records:
    sheet.append([row.get(header, "") for header in headers])
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions
for column, width in {"A": 21, "B": 18, "C": 12, "D": 14, "E": 12, "F": 12, "G": 22, "H": 70, "I": 18, "J": 28}.items():
    sheet.column_dimensions[column].width = width

# Keep one latest regular-stock record per ticker for quick comparison.
latest = {}
for row in records:
    ticker = row["Ticker"]
    if ticker and not any(char.isdigit() for char in ticker):
        latest[ticker] = row
summary = workbook.create_sheet("Latest by Stock")
summary.append(headers)
for cell in summary[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="548235")
for ticker in sorted(latest):
    row = latest[ticker]
    summary.append([row.get(header, "") for header in headers])
summary.freeze_panes = "A2"
summary.auto_filter.ref = summary.dimensions
for column, width in {"A": 21, "B": 18, "C": 12, "D": 14, "E": 12, "F": 12, "G": 22, "H": 70, "I": 18, "J": 28}.items():
    summary.column_dimensions[column].width = width

workbook.save(OUTPUT)
print(f"Date: {payload['date']}")
print(f"Website records in 09:15-15:30 IST: {len(records)}")
print(f"Unique regular-stock tickers: {len(latest)}")
print(f"Saved: {OUTPUT}")
