"""Fetch today's Zee Business website signals and append only new records."""
from __future__ import annotations

import datetime as dt
import os
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill

IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
NOW = dt.datetime.now(IST)
DAY = NOW.strftime("%Y-%m-%d")
OUTPUT = Path(os.environ.get("SIGNAL_XLSX", "zeebiz_live_stocks.xlsx"))
API_URL = "https://www.zeebusinesslive.com/api/trades?day=0"
HEADERS = [
    "Date/Time (IST)", "Ticker", "Position", "Entry Price", "Target",
    "Stop Loss", "Analyst", "Message / Analyst Text", "Category", "Source",
]

if not (dt.time(9, 15) <= NOW.time() <= dt.time(15, 30)):
    print(f"{DAY}: outside market hours; nothing to sync")
    raise SystemExit(0)

def to_ist(value: str | None) -> str:
    if not value:
        return ""
    parsed = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
    return parsed.astimezone(IST).strftime("%Y-%m-%d %H:%M:%S")

def analyst(item: dict) -> str:
    if item.get("sebiAdvisorName"):
        return item["sebiAdvisorName"]
    message = (item.get("message") or "").lower()
    known = {
        "सोमिल मेहता": "Somil Mehta", "सहज अग्रवाल": "Sahaj Aggarwal",
        "विकास सेठी": "Vikas Sethi", "विकास सालुंखे": "Vikas Salunkhe",
        "सुदीप शाह": "Sudeep Shah", "सिद्धार्थ सेडानी": "Siddharth Sedani",
    }
    return next((value for key, value in known.items() if key in message), "")

def load_workbook():
    if OUTPUT.exists():
        workbook = openpyxl.load_workbook(OUTPUT)
        sheet = workbook["Website Signals"] if "Website Signals" in workbook.sheetnames else workbook.active
        sheet.title = "Website Signals"
        return workbook, sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Website Signals"
    sheet.append(HEADERS)
    return workbook, sheet

payload = requests.get(API_URL, timeout=30).json()
if payload.get("date") != DAY:
    raise RuntimeError(f"API returned {payload.get('date')}; local date is {DAY}")

workbook, sheet = load_workbook()
existing = set()
for row in sheet.iter_rows(min_row=2, values_only=True):
    existing.add(tuple(row[:8]))

added = 0
for item in payload.get("data", []):
    stamp = to_ist(item.get("entryOn") or item.get("createdOn"))
    if not stamp.startswith(DAY):
        continue
    time_part = dt.datetime.strptime(stamp, "%Y-%m-%d %H:%M:%S").time()
    if not (dt.time(9, 15) <= time_part <= dt.time(15, 30)):
        continue
    row = (
        stamp, item.get("ticker", ""), item.get("position", ""),
        item.get("entryPrice", ""), item.get("target", ""),
        item.get("stoploss", ""), analyst(item), item.get("message", ""),
        item.get("category", ""), "https://www.zeebusinesslive.com/",
    )
    if tuple(row[:8]) not in existing:
        sheet.append(row)
        existing.add(tuple(row[:8]))
        added += 1

for cell in sheet[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions
for column, width in {"A": 21, "B": 18, "C": 12, "D": 14, "E": 12, "F": 12, "G": 22, "H": 70, "I": 18, "J": 32}.items():
    sheet.column_dimensions[column].width = width

temporary = OUTPUT.with_suffix(OUTPUT.suffix + ".tmp")
workbook.save(temporary)
os.replace(temporary, OUTPUT)
print(f"{DAY}: added {added} new signals; total rows: {sheet.max_row - 1}; output: {OUTPUT}")
