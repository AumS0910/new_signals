"""
Zee Business Live TV → Stock Transcription Pipeline
====================================================
Captures audio from Zee Business Hindi live TV (YouTube stream),
transcribes it using Faster-Whisper (GPU), extracts stock names
and recommendations, and saves to Excel with daily sheets.

Usage:
    python live_transcriber.py              # Run with auto market-hours (9:15 AM - 3:30 PM)
    python live_transcriber.py --no-timer   # Run until manually stopped (Ctrl+C)
    python live_transcriber.py --test       # Quick 2-minute test run

Requirements:
    - ffmpeg (in PATH)
    - yt-dlp (pip install yt-dlp)
    - faster-whisper (pip install faster-whisper)
    - rapidfuzz (pip install rapidfuzz)
    - openpyxl (pip install openpyxl)
    - NVIDIA GPU with CUDA (RTX 3070 Ti or better recommended)
"""

import subprocess
import sys
import os
import io
import re
import time
import signal
import struct
import datetime
import argparse
import threading
import traceback
from pathlib import Path

# Fix Windows terminal encoding for emoji/Unicode
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass  # Fallback: some terminals don't support reconfigure

def safe_print(*args, **kwargs):
    """Print that handles Unicode gracefully on Windows"""
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        text = " ".join(str(a) for a in args)
        # Strip emoji/special chars if encoding fails
        text = text.encode('ascii', 'replace').decode('ascii')
        print(text, **kwargs)



import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz, process

# Import our stock master list
from nse_stocks import STOCK_MASTER_LIST, ACTION_KEYWORDS, PRICE_KEYWORDS, build_lookup


# ============================================================
# Configuration
# ============================================================
CONFIG = {
    # Stream
    "YOUTUBE_CHANNEL": "https://www.youtube.com/@ZeeBusiness/live",
    "YOUTUBE_CHANNEL_ALT": "https://www.youtube.com/@ZeeBusinessHindi/live",

    # Audio
    "SAMPLE_RATE": 16000,       # 16kHz for Whisper
    "CHANNELS": 1,              # Mono
    "CHUNK_DURATION": 30,       # seconds per transcription chunk
    "BYTES_PER_SAMPLE": 2,      # 16-bit PCM = 2 bytes

    # Whisper
    "WHISPER_MODEL": "small",   # small ~2GB VRAM, good balance of speed/accuracy
    "WHISPER_DEVICE": "cuda",   # Use GPU
    "WHISPER_COMPUTE_TYPE": "int8",  # Fastest inference with minimal quality loss
    "WHISPER_LANGUAGE": "hi",   # Hindi (also catches English stock names naturally)
    "WHISPER_BEAM_SIZE": 3,     # Balanced beam search

    # Stock matching
    "FUZZY_THRESHOLD": 82,      # Minimum fuzzy match score (0-100)
    "MIN_WORD_LENGTH": 3,       # Skip very short words for matching
    "RECOMMENDATIONS_ONLY": True, # Ignore ordinary market/news mentions

    # Output
    "EXCEL_FILE": "zeebiz_live_stocks.xlsx",
    "TRANSCRIPT_DIR": "transcripts",
    "AUTOSAVE_INTERVAL": 300,   # Save Excel every 5 minutes (seconds)

    # Market hours (IST)
    "MARKET_OPEN_HOUR": 9,
    "MARKET_OPEN_MINUTE": 15,   # Capture the complete regular market session
    "MARKET_CLOSE_HOUR": 15,
    "MARKET_CLOSE_MINUTE": 30,  # Stop at the regular market close
}


# ============================================================
# Stream Discovery
# ============================================================
class StreamFinder:
    """Find the current Zee Business YouTube live stream URL"""

    @staticmethod
    def find_live_url():
        """Use yt-dlp to get the direct audio stream URL"""
        # The channel's generic /live page can point at a non-market Zee
        # broadcast (for example, breaking-news coverage). Prefer a live
        # result whose title identifies the stock-market programme.
        try:
            result = subprocess.run(
                [
                    "yt-dlp", "--flat-playlist", "--no-warnings",
                    "--print", "%(webpage_url)s\\t%(title)s\\t%(live_status)s",
                    "ytsearch10:Zee Business live stock market",
                ], capture_output=True, text=True, timeout=30
            )
            market_words = ("stock", "share market", "first trade", "bazaar", "anil singhvi", "nifty")
            for line in result.stdout.splitlines():
                parts = line.split("\\t", 2)
                if len(parts) == 3:
                    url, title, status = parts
                    title_lower = title.lower()
                    if status == "is_live" and any(word in title_lower for word in market_words):
                        safe_print(f"✅ Found market stream: {title[:100]}")
                        return url
        except Exception as e:
            safe_print(f"⚠️ Market-stream search failed: {e}")

        urls_to_try = [
            CONFIG["YOUTUBE_CHANNEL"],
            CONFIG["YOUTUBE_CHANNEL_ALT"],
        ]

        for channel_url in urls_to_try:
            safe_print(f"🔍 Searching for live stream at: {channel_url}")
            try:
                # First, get the actual live video URL
                result = subprocess.run(
                    [
                        "yt-dlp",
                        "--no-download",
                        "--print", "webpage_url",
                        "--playlist-items", "1",
                        channel_url,
                    ],
                    capture_output=True, text=True, timeout=30
                )

                if result.returncode == 0 and result.stdout.strip():
                    live_url = result.stdout.strip()
                    safe_print(f"✅ Found live stream: {live_url}")
                    return live_url
                else:
                    safe_print(f"  ❌ No live stream found at {channel_url}")
                    if result.stderr:
                        # Only show first 200 chars of error
                        safe_print(f"  Error: {result.stderr[:200]}")

            except subprocess.TimeoutExpired:
                safe_print(f"  ⏰ Timeout searching {channel_url}")
            except Exception as e:
                safe_print(f"  ❌ Error: {e}")

        return None


# ============================================================
# Audio Capture (yt-dlp download + ffmpeg convert, Windows-safe)
# ============================================================
class AudioCapture:
    """Capture audio from YouTube live stream using yt-dlp download + ffmpeg conversion.
    
    Uses yt-dlp to download audio segments (handles YouTube auth/HLS/DASH),
    then ffmpeg to convert to 16kHz mono WAV for Whisper.
    """

    def __init__(self, stream_url):
        self.stream_url = stream_url
        self._stopped = False
        self._ffmpeg_process = None
        self._temp_dir = os.path.join(os.path.dirname(__file__) or ".", "_audio_tmp")
        os.makedirs(self._temp_dir, exist_ok=True)
        self._chunk_counter = 0

    def start(self):
        """Initialize audio capture"""
        safe_print(f"🎙️ Starting audio capture...")
        safe_print(f"   Chunk duration: {CONFIG['CHUNK_DURATION']}s")
        safe_print(f"   Stream: {self.stream_url}")
        # Resolve once and keep a continuous stream; downloading each 30s
        # section separately cannot keep up with a live broadcast.
        result = subprocess.run(
            ["yt-dlp", "--no-playlist", "--no-warnings", "-f", "bestaudio/best", "--get-url", self.stream_url],
            capture_output=True, text=True, timeout=30,
        )
        direct_url = next((line.strip() for line in result.stdout.splitlines() if line.strip()), None)
        if not direct_url:
            raise RuntimeError(f"Could not resolve audio stream: {result.stderr[:300]}")
        self._ffmpeg_process = subprocess.Popen(
            ["ffmpeg", "-loglevel", "error", "-i", direct_url,
             "-ac", str(CONFIG["CHANNELS"]), "-ar", str(CONFIG["SAMPLE_RATE"]),
             "-f", "s16le", "-acodec", "pcm_s16le", "pipe:1"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, bufsize=0,
        )
        safe_print(f"✅ Audio capture initialized")
        return self

    def read_chunk(self):
        """Download and read one chunk of audio (returns numpy array or None)"""
        if self._ffmpeg_process:
            try:
                expected_bytes = CONFIG["CHUNK_DURATION"] * CONFIG["SAMPLE_RATE"] * CONFIG["CHANNELS"] * CONFIG["BYTES_PER_SAMPLE"]
                safe_print(f"   📥 Reading {CONFIG['CHUNK_DURATION']}s from continuous audio stream...")
                data = self._ffmpeg_process.stdout.read(expected_bytes)
                if len(data) < expected_bytes * 0.5:
                    return None
                audio = np.frombuffer(data, dtype=np.int16).astype(np.float32) / 32768.0
                safe_print(f"   ✅ Audio loaded: {len(audio)/CONFIG['SAMPLE_RATE']:.1f}s")
                return audio
            except Exception as e:
                safe_print(f"   ❌ Audio stream error: {e}")
                return None

        if self._stopped:
            return None

        self._chunk_counter += 1
        raw_file = os.path.join(self._temp_dir, f"chunk_{self._chunk_counter}_raw")
        wav_file = os.path.join(self._temp_dir, f"chunk_{self._chunk_counter}.wav")

        try:
            # Step 1: Use yt-dlp to download a segment of audio
            # --download-sections downloads only the specified time range from live
            # For live streams, we use --match-filter to get current content
            yt_cmd = [
                "yt-dlp",
                "-f", "bestaudio/best",
                "--no-playlist",
                "--no-part",
                "--quiet",
                "--no-warnings",
                "-o", raw_file + ".%(ext)s",
                # For live streams, download a fixed duration
                "--download-sections", f"*0-{CONFIG['CHUNK_DURATION']}",
                self.stream_url,
            ]

            safe_print(f"   📥 Downloading {CONFIG['CHUNK_DURATION']}s audio segment...")
            result = subprocess.run(
                yt_cmd,
                capture_output=True,
                timeout=CONFIG["CHUNK_DURATION"] + 45,
            )

            if result.returncode != 0:
                stderr_text = result.stderr.decode('utf-8', errors='replace')[:300]
                safe_print(f"   ⚠️ yt-dlp error: {stderr_text}")
                
                # Fallback: try without --download-sections (older yt-dlp or different stream type)
                safe_print(f"   🔄 Trying fallback download method...")
                yt_cmd_fallback = [
                    "yt-dlp",
                    "-f", "bestaudio/best",
                    "--no-playlist",
                    "--no-part",
                    "--quiet",
                    "--no-warnings",
                    "--external-downloader", "ffmpeg",
                    "--external-downloader-args", f"ffmpeg:-t {CONFIG['CHUNK_DURATION']}",
                    "-o", raw_file + ".%(ext)s",
                    self.stream_url,
                ]
                result = subprocess.run(
                    yt_cmd_fallback,
                    capture_output=True,
                    timeout=CONFIG["CHUNK_DURATION"] + 60,
                )
                if result.returncode != 0:
                    stderr_text = result.stderr.decode('utf-8', errors='replace')[:300]
                    safe_print(f"   ❌ Fallback also failed: {stderr_text}")
                    return None

            # Find the downloaded file (yt-dlp adds extension)
            downloaded_file = None
            for ext in [".webm", ".m4a", ".opus", ".ogg", ".mp3", ".aac", ".wav", ".mp4"]:
                candidate = raw_file + ext
                if os.path.exists(candidate) and os.path.getsize(candidate) > 100:
                    downloaded_file = candidate
                    break
            
            # Also check for files matching the pattern in temp dir
            if not downloaded_file:
                import glob
                candidates = glob.glob(raw_file + ".*")
                for c in candidates:
                    if os.path.getsize(c) > 100:
                        downloaded_file = c
                        break

            if not downloaded_file:
                safe_print(f"   ⚠️ No audio file found after download")
                return None

            safe_print(f"   📦 Downloaded: {os.path.basename(downloaded_file)} ({os.path.getsize(downloaded_file):,} bytes)")

            # Step 2: Convert to 16kHz mono WAV using ffmpeg
            ffmpeg_cmd = [
                "ffmpeg",
                "-y",
                "-i", downloaded_file,
                "-ac", str(CONFIG["CHANNELS"]),
                "-ar", str(CONFIG["SAMPLE_RATE"]),
                "-f", "wav",
                "-loglevel", "error",
                wav_file,
            ]

            result = subprocess.run(
                ffmpeg_cmd,
                capture_output=True,
                timeout=30,
            )

            if result.returncode != 0:
                stderr_text = result.stderr.decode('utf-8', errors='replace')[:200]
                safe_print(f"   ⚠️ FFmpeg conversion error: {stderr_text}")
                return None

            if not os.path.exists(wav_file) or os.path.getsize(wav_file) < 1000:
                safe_print(f"   ⚠️ WAV file too small or missing")
                return None

            # Step 3: Read WAV file into numpy array
            import wave
            with wave.open(wav_file, 'rb') as wf:
                frames = wf.readframes(wf.getnframes())
                audio = np.frombuffer(frames, dtype=np.int16).astype(np.float32) / 32768.0

            safe_print(f"   ✅ Audio loaded: {len(audio)/CONFIG['SAMPLE_RATE']:.1f}s")
            return audio

        except subprocess.TimeoutExpired:
            safe_print(f"   ⏰ Download timeout — stream may be down")
            return None
        except Exception as e:
            safe_print(f"   ❌ Audio capture error: {e}")
            traceback.print_exc()
            return None
        finally:
            # Clean up temp files for this chunk
            import glob
            for f in glob.glob(raw_file + ".*"):
                try:
                    os.remove(f)
                except:
                    pass
            try:
                if os.path.exists(wav_file):
                    os.remove(wav_file)
            except:
                pass

    def stop(self):
        """Stop audio capture and clean up"""
        self._stopped = True
        if self._ffmpeg_process:
            try:
                self._ffmpeg_process.terminate()
                self._ffmpeg_process.wait(timeout=5)
            except Exception:
                try:
                    self._ffmpeg_process.kill()
                except Exception:
                    pass
            self._ffmpeg_process = None
        try:
            import shutil
            if os.path.exists(self._temp_dir):
                shutil.rmtree(self._temp_dir, ignore_errors=True)
        except:
            pass
        safe_print("🛑 Audio capture stopped")

    def is_alive(self):
        """Check if capture is still active"""
        return not self._stopped





# ============================================================
# Transcription Engine (Faster-Whisper)
# ============================================================
class Transcriber:
    """Hindi speech-to-text using Faster-Whisper"""

    def __init__(self):
        self.model = None

    def load_model(self):
        """Load the Whisper model (downloads on first run)"""
        from faster_whisper import WhisperModel

        safe_print(f"📥 Loading Whisper model '{CONFIG['WHISPER_MODEL']}' on {CONFIG['WHISPER_DEVICE']}...")
        safe_print(f"   Compute type: {CONFIG['WHISPER_COMPUTE_TYPE']}")

        try:
            self.model = WhisperModel(
                CONFIG["WHISPER_MODEL"],
                device=CONFIG["WHISPER_DEVICE"],
                compute_type=CONFIG["WHISPER_COMPUTE_TYPE"],
            )
            safe_print(f"✅ Model loaded successfully!")
        except Exception as e:
            safe_print(f"⚠️ GPU loading failed ({e}), falling back to CPU...")
            CONFIG["WHISPER_DEVICE"] = "cpu"
            CONFIG["WHISPER_COMPUTE_TYPE"] = "int8"
            self.model = WhisperModel(
                CONFIG["WHISPER_MODEL"],
                device="cpu",
                compute_type="int8",
            )
            safe_print(f"✅ Model loaded on CPU (will be slower)")

    def transcribe(self, audio_chunk):
        """Transcribe audio chunk to text"""
        if self.model is None:
            self.load_model()

        try:
            segments, info = self.model.transcribe(
                audio_chunk,
                language=CONFIG["WHISPER_LANGUAGE"],
                beam_size=CONFIG["WHISPER_BEAM_SIZE"],
                vad_filter=True,           # Voice activity detection to skip silence
                vad_parameters=dict(
                    min_silence_duration_ms=500,
                ),
            )

            text_parts = []
            for segment in segments:
                text_parts.append(segment.text.strip())

            full_text = " ".join(text_parts)
            return full_text

        except Exception as e:
            safe_print(f"❌ Transcription error: {e}")
            return ""


# ============================================================
# Stock Extractor
# ============================================================
class StockExtractor:
    """Extract stock names and recommendations from transcribed text"""

    def __init__(self):
        self.symbol_lookup, self.all_aliases = build_lookup()
        # Build a more efficient list for fuzzy matching - only unique items
        self._unique_aliases = list(set(self.all_aliases))
        safe_print(f"📊 Stock extractor loaded: {len(STOCK_MASTER_LIST)} stocks, {len(self._unique_aliases)} unique aliases")

    def extract_stocks(self, text, timestamp=None):
        """Extract stock mentions and recommendations from transcribed text"""
        if not text or len(text.strip()) < 10:
            return []

        results = []
        found_stocks = set()  # Avoid duplicates per chunk

        # Tokenize text into words and multi-word phrases (up to 3 words)
        words = text.split()
        phrases = []
        for i in range(len(words)):
            # Single words
            phrases.append(words[i])
            # Two-word phrases
            if i + 1 < len(words):
                phrases.append(f"{words[i]} {words[i+1]}")
            # Three-word phrases
            if i + 2 < len(words):
                phrases.append(f"{words[i]} {words[i+1]} {words[i+2]}")

        for phrase in phrases:
            clean_phrase = phrase.strip().upper()
            if len(clean_phrase) < CONFIG["MIN_WORD_LENGTH"]:
                continue

            # First try exact match
            if clean_phrase in self.symbol_lookup:
                symbol, full_name = self.symbol_lookup[clean_phrase]
                if symbol not in found_stocks:
                    found_stocks.add(symbol)
                    action = self._detect_action(text, phrase)
                    target, stop_loss = self._extract_prices(text, phrase)
                    analyst = self._detect_analyst(text)
                    if CONFIG["RECOMMENDATIONS_ONLY"] and not self._is_recommendation(text, phrase, action, target, stop_loss):
                        continue
                    results.append({
                        "timestamp": timestamp or datetime.datetime.now().strftime("%H:%M:%S"),
                        "stock": full_name,
                        "symbol": symbol,
                        "action": action,
                        "target": target,
                        "stop_loss": stop_loss,
                        "analyst": analyst,
                        "confidence": 100,
                    })
                continue

            # Fuzzy match for longer phrases (skip very short ones to avoid false positives)
            if len(clean_phrase) >= 4:
                match = process.extractOne(
                    clean_phrase,
                    self._unique_aliases,
                    scorer=fuzz.ratio,
                    score_cutoff=CONFIG["FUZZY_THRESHOLD"],
                )
                if match:
                    matched_alias, score, _ = match
                    if matched_alias in self.symbol_lookup:
                        symbol, full_name = self.symbol_lookup[matched_alias]
                        if symbol not in found_stocks:
                            found_stocks.add(symbol)
                            action = self._detect_action(text, phrase)
                            target, stop_loss = self._extract_prices(text, phrase)
                            analyst = self._detect_analyst(text)
                            if CONFIG["RECOMMENDATIONS_ONLY"] and not self._is_recommendation(text, phrase, action, target, stop_loss):
                                continue
                            results.append({
                                "timestamp": timestamp or datetime.datetime.now().strftime("%H:%M:%S"),
                                "stock": full_name,
                                "symbol": symbol,
                                "action": action,
                                "target": target,
                                "stop_loss": stop_loss,
                                "analyst": analyst,
                                "confidence": int(score),
                            })

        return results

    def _is_recommendation(self, text, stock_phrase, action, target, stop_loss):
        """Keep stock picks, not every stock/index mentioned in commentary."""
        if action or target or stop_loss:
            return True
        position = text.lower().find(stock_phrase.lower())
        if position < 0:
            return False
        window = text[max(0, position - 120):position + len(stock_phrase) + 180]
        return bool(re.search(r"(?:rs\.?|₹|price|level|range|above|below|stop|target|buy|sell|long|short|[0-9]{2,})", window, re.I))

    def _detect_action(self, text, stock_phrase):
        """Detect buy/sell/hold action near the stock mention"""
        text_lower = text.lower()

        # Search in a window around the stock mention
        stock_pos = text_lower.find(stock_phrase.lower())
        if stock_pos >= 0:
            window_start = max(0, stock_pos - 100)
            window_end = min(len(text_lower), stock_pos + len(stock_phrase) + 100)
            window = text_lower[window_start:window_end]
        else:
            window = text_lower

        for action, keywords in ACTION_KEYWORDS.items():
            if action in ("TARGET", "STOP LOSS"):
                continue  # These aren't actions
            for kw in keywords:
                if kw.lower() in window:
                    return action

        return ""

    def _extract_prices(self, text, stock_phrase):
        """Extract target price and stop loss near the stock mention"""
        target = ""
        stop_loss = ""

        # Look for price patterns near the stock name
        text_lower = text.lower()

        # Target patterns
        target_patterns = [
            r'target\s*(?:of\s*)?(?:Rs\.?\s*)?(\d[\d,]*(?:\.\d+)?)',
            r'टारगेट\s*(?:Rs\.?\s*)?(\d[\d,]*(?:\.\d+)?)',
            r'लक्ष्य\s*(?:Rs\.?\s*)?(\d[\d,]*(?:\.\d+)?)',
        ]
        for pattern in target_patterns:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                target = m.group(1).replace(",", "")
                break

        # Stop loss patterns
        sl_patterns = [
            r'stop\s*loss\s*(?:at\s*)?(?:Rs\.?\s*)?(\d[\d,]*(?:\.\d+)?)',
            r'स्टॉप\s*लॉस\s*(?:Rs\.?\s*)?(\d[\d,]*(?:\.\d+)?)',
            r'SL\s*(?:at\s*)?(?:Rs\.?\s*)?(\d[\d,]*(?:\.\d+)?)',
        ]
        for pattern in sl_patterns:
            m = re.search(pattern, text, re.IGNORECASE)
            if m:
                stop_loss = m.group(1).replace(",", "")
                break

        return target, stop_loss

    def _detect_analyst(self, text):
        """Detect analyst names commonly on Zee Business"""
        analysts = [
            ("Somil", ["Somil"]),
            ("Mehul", ["Mehul"]),
            ("Sudip", ["Sudip"]),
            ("Vaishali", ["Vaishali"]),
            ("Sahaj", ["Sahaj"]),
            ("Jay", ["Jay"]),
            ("Sumit", ["Sumit"]),
            ("Anil Singhvi", ["Anil Singhvi", "अनिल सिंघवी", "सिंघवी"]),
            ("Kunal Saraogi", ["Kunal Saraogi", "कुनाल सरावगी", "सरावगी"]),
            ("Rakesh Bansal", ["Rakesh Bansal", "राकेश बंसल", "बंसल"]),
            ("Sandeep Jain", ["Sandeep Jain", "संदीप जैन"]),
            ("Vikas Sethi", ["Vikas Sethi", "विकास सेठी", "सेठी"]),
            ("Ambareesh Baliga", ["Ambareesh Baliga", "अंबरीश बालिगा"]),
            ("Ruchit Jain", ["Ruchit Jain", "रुचित जैन"]),
            ("Varun Aggarwal", ["Varun Aggarwal", "वरुण अग्रवाल"]),
        ]

        text_lower = text.lower()
        for name, aliases in analysts:
            for alias in aliases:
                if re.search(r"(?<![a-z])" + re.escape(alias.lower()) + r"(?![a-z])", text_lower):
                    return name
        return ""


# ============================================================
# Excel Writer
# ============================================================
class ExcelWriter:
    """Write stock data to Excel with daily sheets"""

    def __init__(self, filepath=None):
        self.filepath = filepath or os.path.join(os.path.dirname(__file__), CONFIG["EXCEL_FILE"])
        self.workbook = None
        self._last_save = time.time()
        self._unsaved_changes = False
        self._lock = threading.Lock()

        # Create the workbook and today's sheet immediately.  Previously the
        # file was not created until the first stock match was found, which
        # made it look as if the pipeline was not working.
        self._get_or_create_sheet()
        self.save_if_needed(force=True)

    def _ensure_workbook(self):
        """Load or create the Excel workbook"""
        if self.workbook is None:
            if os.path.exists(self.filepath):
                try:
                    self.workbook = load_workbook(self.filepath)
                    safe_print(f"📂 Loaded existing Excel: {self.filepath}")
                except Exception:
                    self.workbook = Workbook()
                    # Remove default sheet
                    if "Sheet" in self.workbook.sheetnames:
                        del self.workbook["Sheet"]
                    safe_print(f"📝 Created new Excel: {self.filepath}")
            else:
                self.workbook = Workbook()
                if "Sheet" in self.workbook.sheetnames:
                    del self.workbook["Sheet"]
                safe_print(f"📝 Created new Excel: {self.filepath}")

    def _get_or_create_sheet(self, date_str=None):
        """Get or create today's sheet"""
        if date_str is None:
            date_str = datetime.datetime.now().strftime("%Y-%m-%d")

        self._ensure_workbook()

        if date_str in self.workbook.sheetnames:
            return self.workbook[date_str]

        # Create new sheet for today
        ws = self.workbook.create_sheet(title=date_str)

        # Style definitions
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title row
        title_font = Font(name="Calibri", bold=True, color="1F4E79", size=14)
        ws.merge_cells("A1:I1")
        ws["A1"] = f"🔴 Zee Business Live TV — Stock Picks ({date_str})"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Subtitle
        ws.merge_cells("A2:I2")
        ws["A2"] = f"Auto-transcribed from Zee Business Hindi live stream | Market Hours: 9:15 AM – 3:30 PM IST"
        ws["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Headers (row 4)
        headers = [
            "Timestamp", "Stock Name", "NSE Symbol", "Action",
            "Target Price", "Stop Loss", "Analyst", "Confidence %", "Raw Transcript"
        ]
        col_widths = [12, 25, 15, 10, 15, 15, 20, 14, 60]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 9 else None].width = width

        # Set column widths (handle columns A-I)
        for i, width in enumerate(col_widths):
            col_letter = chr(65 + i)  # A=65 in ASCII
            ws.column_dimensions[col_letter].width = width

        # Freeze panes (header stays visible while scrolling)
        ws.freeze_panes = "A5"

        safe_print(f"📋 Created new sheet: {date_str}")
        return ws

    def add_stock_entry(self, entry, raw_transcript=""):
        """Add a stock entry to today's sheet"""
        with self._lock:
            ws = self._get_or_create_sheet()
            next_row = ws.max_row + 1

            # Color coding for actions
            action_colors = {
                "BUY": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),   # Light green
                "SELL": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),   # Light red
                "HOLD": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),   # Light yellow
            }

            action_fonts = {
                "BUY": Font(name="Calibri", bold=True, color="2E7D32"),    # Green
                "SELL": Font(name="Calibri", bold=True, color="C62828"),   # Red
                "HOLD": Font(name="Calibri", bold=True, color="F57F17"),   # Orange
            }

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            values = [
                entry.get("timestamp", ""),
                entry.get("stock", ""),
                entry.get("symbol", ""),
                entry.get("action", ""),
                entry.get("target", ""),
                entry.get("stop_loss", ""),
                entry.get("analyst", ""),
                entry.get("confidence", ""),
                raw_transcript[:200] if raw_transcript else "",  # Truncate long text
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 9))

                # Apply action coloring
                action = entry.get("action", "")
                if action in action_colors:
                    cell.fill = action_colors[action]
                    if col_idx == 4:  # Action column
                        cell.font = action_fonts.get(action, Font(name="Calibri"))

            self._unsaved_changes = True

        # Persist each detected stock immediately.  The periodic autosave in
        # the main loop remains as a recovery measure for any later changes.
        self.save_if_needed(force=True)

    def save_if_needed(self, force=False):
        """Save workbook if there are unsaved changes"""
        with self._lock:
            if not self._unsaved_changes and not force:
                return

            elapsed = time.time() - self._last_save
            if not force and elapsed < CONFIG["AUTOSAVE_INTERVAL"]:
                return

            if self.workbook:
                try:
                    self.workbook.save(self.filepath)
                    self._last_save = time.time()
                    self._unsaved_changes = False
                    safe_print(f"💾 Excel saved: {self.filepath}")
                except PermissionError:
                    safe_print(f"⚠️ Cannot save — file is open in Excel. Will retry later.")
                except Exception as e:
                    safe_print(f"❌ Save error: {e}")

    def force_save(self):
        """Force save (used during shutdown)"""
        self.save_if_needed(force=True)


# ============================================================
# Transcript Logger
# ============================================================
class TranscriptLogger:
    """Save full transcript text to daily log files"""

    def __init__(self):
        self.transcript_dir = os.path.join(os.path.dirname(__file__), CONFIG["TRANSCRIPT_DIR"])
        os.makedirs(self.transcript_dir, exist_ok=True)

    def log(self, text, timestamp=None):
        """Append transcribed text to today's log file"""
        if not text or not text.strip():
            return

        date_str = datetime.datetime.now().strftime("%Y-%m-%d")
        ts = timestamp or datetime.datetime.now().strftime("%H:%M:%S")
        filepath = os.path.join(self.transcript_dir, f"{date_str}.txt")

        try:
            with open(filepath, "a", encoding="utf-8") as f:
                f.write(f"[{ts}] {text.strip()}\n")
        except Exception as e:
            safe_print(f"❌ Transcript log error: {e}")


# ============================================================
# Market Hours Checker
# ============================================================
class MarketHours:
    """Check if we're within Indian stock market hours"""

    @staticmethod
    def is_market_open():
        """Check if current time is within market hours (IST)"""
        # Get IST time (UTC+5:30)
        now = datetime.datetime.now()
        market_open = now.replace(
            hour=CONFIG["MARKET_OPEN_HOUR"],
            minute=CONFIG["MARKET_OPEN_MINUTE"],
            second=0
        )
        market_close = now.replace(
            hour=CONFIG["MARKET_CLOSE_HOUR"],
            minute=CONFIG["MARKET_CLOSE_MINUTE"],
            second=0
        )
        return market_open <= now <= market_close

    @staticmethod
    def is_weekday():
        """Check if today is a weekday (Mon-Fri)"""
        return datetime.datetime.now().weekday() < 5  # 0=Mon, 4=Fri

    @staticmethod
    def time_until_market_open():
        """Seconds until market opens"""
        now = datetime.datetime.now()
        market_open = now.replace(
            hour=CONFIG["MARKET_OPEN_HOUR"],
            minute=CONFIG["MARKET_OPEN_MINUTE"],
            second=0
        )
        if now >= market_open:
            # Market already opened today, check tomorrow
            market_open += datetime.timedelta(days=1)
        delta = (market_open - now).total_seconds()
        return max(0, delta)

    @staticmethod
    def time_until_market_close():
        """Seconds until market closes"""
        now = datetime.datetime.now()
        market_close = now.replace(
            hour=CONFIG["MARKET_CLOSE_HOUR"],
            minute=CONFIG["MARKET_CLOSE_MINUTE"],
            second=0
        )
        delta = (market_close - now).total_seconds()
        return max(0, delta)


# ============================================================
# Main Pipeline
# ============================================================
class LiveTranscriptionPipeline:
    """Main pipeline orchestrator"""

    def __init__(self, no_timer=False, test_mode=False):
        self.no_timer = no_timer
        self.test_mode = test_mode
        self.running = False
        self.audio_capture = None

        # Components
        self.transcriber = Transcriber()
        self.stock_extractor = StockExtractor()
        self.excel_writer = ExcelWriter()
        self.transcript_logger = TranscriptLogger()

        # Stats
        self.chunks_processed = 0
        self.stocks_found = 0
        self.start_time = None

    def _handle_signal(self, signum, frame):
        """Graceful shutdown on Ctrl+C"""
        safe_print("\n\n🛑 Shutdown requested (Ctrl+C)...")
        self.running = False

    def _print_banner(self):
        """Print startup banner"""
        safe_print("=" * 70)
        safe_print("  📺 Zee Business Live TV → Stock Transcription Pipeline")
        safe_print("=" * 70)
        safe_print(f"  Model     : Faster-Whisper ({CONFIG['WHISPER_MODEL']}) on {CONFIG['WHISPER_DEVICE']}")
        safe_print(f"  Language   : Hindi (hi)")
        safe_print(f"  Chunk      : {CONFIG['CHUNK_DURATION']}s")
        safe_print(f"  Excel      : {CONFIG['EXCEL_FILE']}")
        safe_print(f"  Transcripts: {CONFIG['TRANSCRIPT_DIR']}/")
        if self.test_mode:
            safe_print(f"  Mode       : ⚡ TEST (2 minutes)")
        elif self.no_timer:
            safe_print(f"  Mode       : 🔄 Manual (until Ctrl+C)")
        else:
            safe_print(f"  Mode       : ⏰ Market Hours (9:{CONFIG['MARKET_OPEN_MINUTE']:02d} AM – {CONFIG['MARKET_CLOSE_HOUR']-12}:{CONFIG['MARKET_CLOSE_MINUTE']:02d} PM)")
        safe_print("=" * 70)

    def _wait_for_market(self):
        """Wait for market hours to begin"""
        if self.no_timer or self.test_mode:
            return True

        if not MarketHours.is_weekday():
            safe_print("📅 Today is a weekend. Market is closed.")
            safe_print("   The script will wait for the next trading day.")
            # Wait until Monday
            while not MarketHours.is_weekday() and self.running:
                time.sleep(60)

        if not MarketHours.is_market_open():
            wait_time = MarketHours.time_until_market_open()
            if wait_time > 0:
                hours = int(wait_time // 3600)
                minutes = int((wait_time % 3600) // 60)
                safe_print(f"⏳ Market opens in {hours}h {minutes}m. Waiting...")
                while wait_time > 0 and self.running:
                    time.sleep(min(60, wait_time))
                    wait_time = MarketHours.time_until_market_open()

        return self.running

    def _should_continue(self):
        """Check if we should continue running"""
        if not self.running:
            return False

        if self.test_mode:
            elapsed = time.time() - self.start_time
            if elapsed >= 120:  # 2 minutes test
                safe_print("⚡ Test mode complete (2 minutes)")
                return False
            return True

        if not self.no_timer:
            if not MarketHours.is_market_open():
                safe_print("🔔 Market hours ended. Stopping capture.")
                return False

        return True

    def run(self):
        """Main pipeline execution loop"""
        self.running = True
        self.start_time = time.time()

        # Setup signal handler for graceful shutdown
        signal.signal(signal.SIGINT, self._handle_signal)
        signal.signal(signal.SIGTERM, self._handle_signal)

        self._print_banner()

        # Wait for market hours
        if not self._wait_for_market():
            return

        # Load Whisper model (may take a moment on first run to download)
        safe_print("\n📦 Initializing components...")
        self.transcriber.load_model()

        # Main retry loop
        reconnect_attempts = 0
        max_reconnect = 10

        while self.running and self._should_continue() and reconnect_attempts < max_reconnect:
            try:
                # Find live stream
                stream_url = StreamFinder.find_live_url()
                if not stream_url:
                    reconnect_attempts += 1
                    wait = min(60 * reconnect_attempts, 300)  # Back off up to 5 min
                    safe_print(f"⏳ No stream found. Retry {reconnect_attempts}/{max_reconnect} in {wait}s...")
                    time.sleep(wait)
                    continue

                reconnect_attempts = 0  # Reset on successful find

                # Start audio capture
                self.audio_capture = AudioCapture(stream_url)
                self.audio_capture.start()

                # Processing loop
                consecutive_empty = 0
                while self._should_continue():
                    # Read audio chunk
                    audio = self.audio_capture.read_chunk()

                    if audio is None:
                        consecutive_empty += 1
                        if consecutive_empty >= 3:
                            safe_print("⚠️ Stream appears dead. Reconnecting...")
                            break
                        time.sleep(2)
                        continue

                    consecutive_empty = 0
                    self.chunks_processed += 1
                    timestamp = datetime.datetime.now().strftime("%H:%M:%S")

                    # Transcribe
                    safe_print(f"\n🎯 [{timestamp}] Transcribing chunk #{self.chunks_processed}...")
                    text = self.transcriber.transcribe(audio)

                    if text and text.strip():
                        # Log full transcript
                        self.transcript_logger.log(text, timestamp)
                        safe_print(f"   📝 \"{text[:120]}{'...' if len(text) > 120 else ''}\"")

                        # Extract stocks
                        stocks = self.stock_extractor.extract_stocks(text, timestamp)

                        if stocks:
                            for stock in stocks:
                                self.stocks_found += 1
                                action_emoji = {"BUY": "🟢", "SELL": "🔴", "HOLD": "🟡"}.get(stock["action"], "⚪")
                                safe_print(f"   {action_emoji} {stock['symbol']} ({stock['stock']}) "
                                      f"— {stock['action'] or 'Mentioned'}"
                                      f"{' | Target: ₹' + stock['target'] if stock['target'] else ''}"
                                      f"{' | SL: ₹' + stock['stop_loss'] if stock['stop_loss'] else ''}"
                                      f"{' | By: ' + stock['analyst'] if stock['analyst'] else ''}"
                                      f" [{stock['confidence']}%]")

                                # Write to Excel
                                self.excel_writer.add_stock_entry(stock, raw_transcript=text)
                        else:
                            safe_print(f"   (no stock mentions detected)")
                    else:
                        safe_print(f"   🔇 (silence or no speech detected)")

                    # Periodic auto-save
                    self.excel_writer.save_if_needed()

                # Clean up audio capture before reconnect
                if self.audio_capture:
                    self.audio_capture.stop()

            except Exception as e:
                safe_print(f"\n❌ Pipeline error: {e}")
                traceback.print_exc()
                reconnect_attempts += 1
                if self.audio_capture:
                    self.audio_capture.stop()
                time.sleep(10)

        # Shutdown
        self._shutdown()

    def _shutdown(self):
        """Clean shutdown"""
        safe_print("\n" + "=" * 70)
        safe_print("  🛑 Shutting down...")
        safe_print("=" * 70)

        # Stop audio
        if self.audio_capture:
            self.audio_capture.stop()

        # Force save Excel
        self.excel_writer.force_save()

        # Print stats
        elapsed = time.time() - (self.start_time or time.time())
        hours = int(elapsed // 3600)
        minutes = int((elapsed % 3600) // 60)
        safe_print(f"\n📊 Session Summary:")
        safe_print(f"   Duration       : {hours}h {minutes}m")
        safe_print(f"   Chunks processed: {self.chunks_processed}")
        safe_print(f"   Stocks found    : {self.stocks_found}")
        safe_print(f"   Excel file      : {self.excel_writer.filepath}")
        transcript_dir = os.path.join(os.path.dirname(__file__), CONFIG["TRANSCRIPT_DIR"])
        safe_print(f"   Transcripts     : {transcript_dir}/")
        safe_print("=" * 70)
        safe_print("  ✅ All data saved. Goodbye!")
        safe_print("=" * 70)


# ============================================================
# Entry Point
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Zee Business Live TV → Stock Transcription Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python live_transcriber.py              # Run during market hours (9:15 AM - 3:30 PM)
  python live_transcriber.py --no-timer   # Run until Ctrl+C
  python live_transcriber.py --test       # Quick 2-minute test
  python live_transcriber.py --model medium  # Use larger model (more accurate, slower)
        """
    )
    parser.add_argument("--no-timer", action="store_true",
                       help="Run until manually stopped (ignore market hours)")
    parser.add_argument("--test", action="store_true",
                       help="Quick 2-minute test run")
    parser.add_argument("--model", type=str, default="small",
                       choices=["tiny", "base", "small", "medium", "large-v3"],
                       help="Whisper model size (default: small)")
    parser.add_argument("--cpu", action="store_true",
                       help="Force CPU mode (no GPU)")
    parser.add_argument("--chunk", type=int, default=30,
                       help="Audio chunk duration in seconds (default: 30)")

    args = parser.parse_args()

    # Apply config overrides
    CONFIG["WHISPER_MODEL"] = args.model
    CONFIG["CHUNK_DURATION"] = args.chunk
    if args.cpu:
        CONFIG["WHISPER_DEVICE"] = "cpu"
        CONFIG["WHISPER_COMPUTE_TYPE"] = "int8"

    # Run pipeline
    pipeline = LiveTranscriptionPipeline(
        no_timer=args.no_timer,
        test_mode=args.test,
    )
    pipeline.run()


if __name__ == "__main__":
    main()
